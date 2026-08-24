from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from html import escape

from g_file_studio.services.html_report_selection import selection_bar, selection_cell, selection_header, selection_script, selection_style
from g_file_studio.services.report_i18n import report_is_english, report_text
from pathlib import Path
from typing import Iterable


_NAME_ALIASES = {"rmu名称", "rmuid", "rmu id", "rmu name", "name", "名称", "环网柜名称", "环网柜id", "柜名"}
_TYPE_ALIASES = {"rmu类型", "rmu type", "type", "类型", "柜型", "环网柜类型"}
_SMART_ALIASES = {"是否智能", "智能环网柜", "smart", "intelligent", "is smart", "智能", "是否smart"}


@dataclass(slots=True)
class LedgerRow:
    name: str
    rmu_type: str = ""
    intelligent: str = ""  # "YES" / "NO" / ""
    source_row: int = 0


@dataclass(slots=True)
class GraphicRmuRow:
    file_name: str
    name: str
    rmu_type: str
    intelligent: str
    intelligent_source: str
    confidence: str
    duplicate: str
    rect_id: str
    rect_x: str
    rect_y: str
    rect_w: str
    rect_h: str
    warnings: str


@dataclass(slots=True)
class ComparisonRow:
    ledger_name: str
    graphic_name: str
    ledger_type: str
    graphic_type: str
    ledger_intelligent: str
    graphic_intelligent: str
    intelligent_source: str
    file_name: str
    rect_id: str
    confidence: str
    graphic_duplicate: str
    ledger_duplicate: str
    result: str
    warnings: str


def _norm_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().casefold())


def _norm_name(value: object) -> str:
    return str(value or "").strip()


def _norm_type(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def _norm_intelligent(value: object) -> str:
    text = str(value or "").strip().casefold()
    if not text:
        return ""
    yes = {"1", "yes", "y", "true", "是", "智能", "smart", "smr"}
    no = {"0", "no", "n", "false", "否", "普通", "non-smart", "nonsmart"}
    if text in yes:
        return "YES"
    if text in no:
        return "NO"
    raise ValueError(f"无法识别‘是否智能’值：{value}")


def _header_index(headers: list[str], aliases: set[str]) -> int | None:
    normalized_aliases = {_norm_header(v) for v in aliases}
    for index, header in enumerate(headers):
        if _norm_header(header) in normalized_aliases:
            return index
    return None


def _rows_from_matrix(matrix: list[list[object]]) -> list[LedgerRow]:
    matrix = [list(row) for row in matrix if any(str(v or "").strip() for v in row)]
    if not matrix:
        return []
    headers = [str(v or "").strip() for v in matrix[0]]
    name_idx = _header_index(headers, _NAME_ALIASES)
    type_idx = _header_index(headers, _TYPE_ALIASES)
    smart_idx = _header_index(headers, _SMART_ALIASES)
    start = 1
    if name_idx is None:
        # 没有表头时按第一列名称、第二列类型、第三列智能处理。
        name_idx, type_idx, smart_idx, start = 0, 1, 2, 0
    rows: list[LedgerRow] = []
    for row_no, row in enumerate(matrix[start:], start=start + 1):
        name = _norm_name(row[name_idx] if name_idx < len(row) else "")
        if not name:
            continue
        rmu_type = _norm_type(row[type_idx] if type_idx is not None and type_idx < len(row) else "")
        intelligent = _norm_intelligent(row[smart_idx] if smart_idx is not None and smart_idx < len(row) else "")
        rows.append(LedgerRow(name=name, rmu_type=rmu_type, intelligent=intelligent, source_row=row_no))
    return rows


def load_ledger_file(path: Path) -> list[LedgerRow]:
    path = Path(path)
    if not path.is_file():
        raise ValueError(f"RMU 台账文件不存在：{path}")
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("读取 Excel 台账需要安装 openpyxl。") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        matrix = [list(row) for row in sheet.iter_rows(values_only=True)]
        workbook.close()
        return _rows_from_matrix(matrix)
    if suffix == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                with path.open("r", encoding=encoding, newline="") as stream:
                    return _rows_from_matrix([list(row) for row in csv.reader(stream)])
            except UnicodeDecodeError as exc:
                last_error = exc
        raise ValueError(f"CSV 编码无法识别：{path}") from last_error
    raise ValueError("RMU 台账仅支持 .xlsx / .xlsm / .csv。")


def parse_pasted_table(text: str) -> list[LedgerRow]:
    lines = [line for line in text.splitlines() if line.strip()]
    if not lines:
        return []
    matrix: list[list[object]] = []
    for line in lines:
        if "\t" in line:
            matrix.append([part.strip() for part in line.split("\t")])
        elif "," in line:
            matrix.append(next(csv.reader([line])))
        else:
            matrix.append(re.split(r"\s{2,}", line.strip()))
    return _rows_from_matrix(matrix)


def parse_name_list(text: str) -> list[LedgerRow]:
    rows: list[LedgerRow] = []
    for row_no, line in enumerate(text.splitlines(), 1):
        value = line.strip().strip(",")
        if not value:
            continue
        if "\t" in value:
            value = value.split("\t", 1)[0].strip()
        elif "," in value:
            value = value.split(",", 1)[0].strip()
        if value:
            rows.append(LedgerRow(name=value, source_row=row_no))
    return rows


def compare_ledger(ledger_rows: list[LedgerRow], graphic_rows: list[GraphicRmuRow]) -> tuple[list[ComparisonRow], dict[str, int]]:
    ledger_counts: dict[str, int] = {}
    graphic_counts: dict[str, int] = {}
    for row in ledger_rows:
        key = row.name.casefold()
        ledger_counts[key] = ledger_counts.get(key, 0) + 1
    for row in graphic_rows:
        key = row.name.strip().casefold()
        if key:
            graphic_counts[key] = graphic_counts.get(key, 0) + 1

    ledger_by_name: dict[str, LedgerRow] = {}
    for row in ledger_rows:
        ledger_by_name.setdefault(row.name.casefold(), row)
    graphic_by_name: dict[str, GraphicRmuRow] = {}
    unnamed_graphics: list[GraphicRmuRow] = []
    for row in graphic_rows:
        key = row.name.strip().casefold()
        if key:
            graphic_by_name.setdefault(key, row)
        else:
            unnamed_graphics.append(row)

    all_keys = sorted(set(ledger_by_name) | set(graphic_by_name))
    output: list[ComparisonRow] = []
    for key in all_keys:
        ledger = ledger_by_name.get(key)
        graphic = graphic_by_name.get(key)
        ledger_dup = ledger_counts.get(key, 0) > 1
        graphic_dup = graphic_counts.get(key, 0) > 1
        result_parts: list[str] = []
        warnings: list[str] = []
        if ledger is None:
            result_parts.append("台账缺失")
        elif graphic is None:
            result_parts.append("图形缺失")
        else:
            if ledger.rmu_type and _norm_type(graphic.rmu_type) != ledger.rmu_type:
                result_parts.append("柜型不一致")
            if ledger.intelligent and graphic.intelligent != ledger.intelligent:
                result_parts.append("智能属性不一致")
            if not result_parts:
                result_parts.append("完全一致")
        if ledger_dup:
            result_parts.append("台账名称重复")
        if graphic_dup:
            result_parts.append("图形名称重复")
        if graphic and graphic.warnings:
            warnings.append(graphic.warnings)
        output.append(ComparisonRow(
            ledger_name=ledger.name if ledger else "",
            graphic_name=graphic.name if graphic else "",
            ledger_type=ledger.rmu_type if ledger else "",
            graphic_type=graphic.rmu_type if graphic else "",
            ledger_intelligent=ledger.intelligent if ledger else "",
            graphic_intelligent=graphic.intelligent if graphic else "",
            intelligent_source=graphic.intelligent_source if graphic else "",
            file_name=graphic.file_name if graphic else "",
            rect_id=graphic.rect_id if graphic else "",
            confidence=graphic.confidence if graphic else "",
            graphic_duplicate="YES" if graphic_dup else "NO",
            ledger_duplicate="YES" if ledger_dup else "NO",
            result="；".join(result_parts),
            warnings=" | ".join(warnings),
        ))

    for graphic in unnamed_graphics:
        output.append(ComparisonRow(
            ledger_name="", graphic_name="", ledger_type="", graphic_type=graphic.rmu_type,
            ledger_intelligent="", graphic_intelligent=graphic.intelligent,
            intelligent_source=graphic.intelligent_source, file_name=graphic.file_name,
            rect_id=graphic.rect_id, confidence=graphic.confidence,
            graphic_duplicate="NO", ledger_duplicate="NO", result="图形柜名未识别",
            warnings=graphic.warnings,
        ))

    stats = {
        "ledger_count": len(ledger_rows),
        "graphic_count": len(graphic_rows),
        "matched_count": sum(1 for row in output if row.result == "完全一致"),
        "type_mismatch_count": sum(1 for row in output if "柜型不一致" in row.result),
        "intelligent_mismatch_count": sum(1 for row in output if "智能属性不一致" in row.result),
        "graphic_missing_count": sum(1 for row in output if "图形缺失" in row.result),
        "ledger_missing_count": sum(1 for row in output if "台账缺失" in row.result),
        "graphic_duplicate_count": sum(1 for key, count in graphic_counts.items() if count > 1),
        "ledger_duplicate_count": sum(1 for key, count in ledger_counts.items() if count > 1),
        "graphic_unnamed_count": len(unnamed_graphics),
    }
    return output, stats


def write_comparison_reports(output_dir: Path, rows: list[ComparisonRow], stats: dict[str, int]) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "rmu-ledger-comparison.csv"
    html_path = output_dir / "rmu-ledger-comparison.html"
    headers = [
        "LedgerRMUName", "GraphicRMUName", "LedgerType", "GraphicType", "LedgerIntelligent",
        "GraphicIntelligent", "IntelligentSource", "File", "RectID", "Confidence",
        "GraphicDuplicate", "LedgerDuplicate", "Result", "Warnings",
    ]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([
                row.ledger_name, row.graphic_name, row.ledger_type, row.graphic_type,
                row.ledger_intelligent, row.graphic_intelligent, row.intelligent_source,
                row.file_name, row.rect_id, report_text(row.confidence), row.graphic_duplicate,
                row.ledger_duplicate, report_text(row.result), report_text(row.warnings),
            ])

    def css_class(result: str) -> str:
        if result == "完全一致":
            return "good"
        if "台账缺失" in result:
            return "warn"
        return "bad"

    body: list[str] = []
    for row in rows:
        values = [
            row.ledger_name, row.graphic_name, row.ledger_type, row.graphic_type,
            row.ledger_intelligent, row.graphic_intelligent, row.intelligent_source,
            row.file_name, row.rect_id, report_text(row.confidence), row.graphic_duplicate,
            row.ledger_duplicate, report_text(row.result), report_text(row.warnings),
        ]
        body.append(
            f'<tr class="{css_class(row.result)}">' + selection_cell()
            + ''.join(f'<td>{escape(str(v))}</td>' for v in values) + '</tr>'
        )

    english = report_is_english()
    summary_labels = ([
        ("ledger_count", "Ledger RMUs"), ("graphic_count", "Graphic RMUs"), ("matched_count", "Fully Matched"),
        ("type_mismatch_count", "Cabinet Type Mismatch"), ("intelligent_mismatch_count", "Smart Attribute Mismatch"),
        ("graphic_missing_count", "Missing in Graphic"), ("ledger_missing_count", "Missing in Ledger"),
        ("graphic_duplicate_count", "Duplicate Graphic Names"), ("ledger_duplicate_count", "Duplicate Ledger Names"),
        ("graphic_unnamed_count", "Unrecognized Graphic RMU Names"),
    ] if english else [
        ("ledger_count", "台账 RMU"), ("graphic_count", "图形 RMU"), ("matched_count", "完全一致"),
        ("type_mismatch_count", "柜型不一致"), ("intelligent_mismatch_count", "智能属性不一致"),
        ("graphic_missing_count", "图形缺失"), ("ledger_missing_count", "台账缺失"),
        ("graphic_duplicate_count", "图形重复名称"), ("ledger_duplicate_count", "台账重复名称"),
        ("graphic_unnamed_count", "图形柜名未识别"),
    ])
    separator = ": " if english else "："
    summary = ''.join(
        f'<div class="badge">{escape(label)}{separator}<b>{stats[key]}</b></div>'
        for key, label in summary_labels
    )
    report_title = "RMU Ledger Comparison Report" if english else "RMU 台账对比报告"
    report_heading = "RMU Ledger vs. G-File Graphic Comparison Report" if english else "RMU 台账与 G 图形对比报告"
    report_legend = (
        "Green = fully matched; yellow = present in the G graphic but missing from the ledger; "
        "red = missing in graphic, cabinet-type/smart-attribute mismatch, duplicate, or unrecognized RMU name."
        if english else
        "绿色=完全一致；黄色=G 图形中存在但台账缺失；红色=图形缺失、柜型/智能属性不一致、重复或柜名未识别。"
    )
    html = f'''<!DOCTYPE html>
<html lang="{'en' if english else 'zh-CN'}"><head><meta charset="utf-8"><title>{report_title}</title>
<style>
body{{font-family:"Microsoft YaHei",Arial,sans-serif;margin:24px;color:#1f2937}}h1{{font-size:24px}}
.summary{{display:flex;flex-wrap:wrap;gap:10px;margin:14px 0}}.badge{{border:1px solid #d1d5db;border-radius:6px;padding:7px 10px;background:#f8fafc}}
table{{width:100%;border-collapse:collapse;font-size:13px}}th,td{{border:1px solid #cbd5e1;padding:6px 8px;text-align:left;vertical-align:top}}th{{background:#e2e8f0;position:sticky;top:0}}
tr.good td{{background:#dcfce7}}tr.warn td{{background:#fef9c3}}tr.bad td{{background:#fee2e2}}
{selection_style()}
</style></head><body><h1>{report_heading}</h1><div class="summary">{summary}</div>
<p>{report_legend}</p>
{selection_bar()}
<table><thead><tr>{selection_header()}{''.join(f'<th>{escape(h)}</th>' for h in headers)}</tr></thead><tbody>{''.join(body)}</tbody></table>{selection_script()}</body></html>'''
    html_path.write_text(html, encoding="utf-8")
    return csv_path, html_path
